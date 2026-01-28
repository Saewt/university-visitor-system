from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, timedelta
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..database import get_db, turkey_now
from ..models import Student, Department
from ..schemas import StudentList
from ..routers.auth import require_admin

router = APIRouter()


def create_excel_file(students: list, summary: dict):
    """Create an Excel file with student data and summary statistics"""
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # --- Student Data Sheet ---
    ws_data = wb.create_sheet("Öğrenci Kayıtları", 0)

    # Headers
    headers = [
        "ID", "Ad", "Soyad", "E-posta", "Telefon",
        "Lise", "Sıralama", "YKS Puanı", "YKS Türü",
        "Bölüm", "Tur İsteği", "Kayıt Tarihi"
    ]

    # Header styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    data_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    row_num = 2

    for student in students:
        ws_data.cell(row=row_num, column=1).value = student.id
        ws_data.cell(row=row_num, column=2).value = student.first_name
        ws_data.cell(row=row_num, column=3).value = student.last_name
        ws_data.cell(row=row_num, column=4).value = student.email
        ws_data.cell(row=row_num, column=5).value = student.phone
        ws_data.cell(row=row_num, column=6).value = student.high_school
        ws_data.cell(row=row_num, column=7).value = student.ranking
        ws_data.cell(row=row_num, column=8).value = float(student.yks_score) if student.yks_score else ""
        ws_data.cell(row=row_num, column=9).value = student.yks_type
        ws_data.cell(row=row_num, column=10).value = student.department_name
        ws_data.cell(row=row_num, column=11).value = "Evet" if student.wants_tour else "Hayır"
        ws_data.cell(row=row_num, column=12).value = student.created_at.strftime("%d.%m.%Y %H:%M")

        # Apply borders and alternating fill
        for col_num in range(1, 13):
            cell = ws_data.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if row_num % 2 == 0:
                cell.fill = data_fill

        row_num += 1

    # Auto-adjust column widths
    for col_num in range(1, 13):
        column_letter = get_column_letter(col_num)
        ws_data.column_dimensions[column_letter].width = 15

    # Freeze header row
    ws_data.freeze_panes = "A2"

    # --- Summary Sheet ---
    ws_summary = wb.create_sheet("Özet İstatistikler", 1)

    summary_data = [
        ["Rapor Tarihi", turkey_now().strftime("%d.%m.%Y %H:%M")],
        [],
        ["Toplam Öğrenci Sayısı", summary.get("total_students", 0)],
        ["Bugünkü Kayıtlar", summary.get("today_count", 0)],
        ["Tur İsteği", summary.get("tour_requests", 0)],
        [],
        ["Bölüm Dağılımı", ""],
    ]

    for dept_stat in summary.get("by_department", []):
        summary_data.append([
            f"  - {dept_stat['department_name']}",
            dept_stat["count"]
        ])

    summary_data.append([])
    summary_data.append(["YKS Türü Dağılımı", ""])

    type_mapping = {
        "SAYISAL": "Sayısal",
        "SOZEL": "Sözel",
        "EA": "Eşit Ağırlık",
        "DIL": "Dil"
    }

    for type_stat in summary.get("by_type", []):
        summary_data.append([
            f"  - {type_mapping.get(type_stat['yks_type'], type_stat['yks_type'])}",
            type_stat["count"]
        ])

    # Write summary data
    for row_num, row_data in enumerate(summary_data, 1):
        if not row_data:  # Skip empty rows
            continue

        label, value = row_data[0], row_data[1] if len(row_data) > 1 else ""
        ws_summary.cell(row=row_num, column=1).value = label
        ws_summary.cell(row=row_num, column=2).value = value

        # Apply styles
        cell1 = ws_summary.cell(row=row_num, column=1)
        cell2 = ws_summary.cell(row=row_num, column=2)

        if row_num <= 2 or label == "Bölüm Dağılımı" or label == "YKS Türü Dağılımı":
            cell1.font = Font(bold=True, size=12)
            cell1.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

        cell2.alignment = Alignment(horizontal="right")

    ws_summary.column_dimensions["A"].width = 35
    ws_summary.column_dimensions["B"].width = 15

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def _get_export_data(db: Session, start_date: Optional[datetime] = None,
                    end_date: Optional[datetime] = None,
                    department_id: Optional[int] = None):
    """Helper function to get export data"""
    from sqlalchemy import func, desc

    query = db.query(
        Student.id,
        Student.first_name,
        Student.last_name,
        Student.email,
        Student.phone,
        Student.high_school,
        Student.ranking,
        Student.yks_score,
        Student.yks_type,
        Student.wants_tour,
        Student.department_id,
        Department.name.label("department_name"),
        Student.created_at
    ).outerjoin(Department)

    # Apply filters
    if start_date:
        query = query.filter(Student.created_at >= start_date)

    if end_date:
        query = query.filter(Student.created_at <= end_date)

    if department_id:
        query = query.filter(Student.department_id == department_id)

    query = query.order_by(Student.created_at.desc())
    students = query.all()

    # Convert to StudentList objects
    student_list = [
        StudentList(
            id=s.id,
            first_name=s.first_name,
            last_name=s.last_name,
            email=s.email,
            phone=s.phone,
            high_school=s.high_school,
            ranking=s.ranking,
            yks_score=float(s.yks_score) if s.yks_score else None,
            yks_type=s.yks_type,
            wants_tour=s.wants_tour,
            department_id=s.department_id,
            department_name=s.department_name,
            created_at=s.created_at
        )
        for s in students
    ]

    # Get summary stats
    query = db.query(Student)

    if start_date:
        query = query.filter(Student.created_at >= start_date)
    if end_date:
        query = query.filter(Student.created_at <= end_date)
    if department_id:
        query = query.filter(Student.department_id == department_id)

    total_students = query.count()

    # Today's count
    today_start = turkey_now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_count = db.query(func.count(Student.id)).filter(
        Student.created_at >= today_start
    ).scalar() or 0

    # Tour requests
    tour_requests = query.filter(Student.wants_tour == True).count()

    # By department
    dept_results = db.query(
        Department.name,
        func.count(Student.id).label("count")
    ).outerjoin(Student, Department.id == Student.department_id).filter(
        Student.created_at >= start_date if start_date else True,
        Student.created_at <= end_date if end_date else True
    ).group_by(Department.id, Department.name).order_by(
        desc("count")
    ).all()

    by_department = [
        {"department_name": row.name or "Belirtilmemiş", "count": row.count or 0}
        for row in dept_results
    ]

    # By YKS type
    type_results = db.query(
        Student.yks_type,
        func.count(Student.id).label("count")
    ).filter(
        Student.yks_type.isnot(None),
        Student.created_at >= start_date if start_date else True,
        Student.created_at <= end_date if end_date else True
    ).group_by(Student.yks_type).all()

    by_type = [
        {"yks_type": row.yks_type, "count": row.count}
        for row in type_results
    ]

    summary = {
        "total_students": total_students,
        "today_count": today_count,
        "tour_requests": tour_requests,
        "by_department": by_department,
        "by_type": by_type
    }

    return student_list, summary


@router.get("/excel")
async def export_excel(
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    department_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user = Depends(require_admin)
):
    """Export student data to Excel with optional filters"""

    student_list, summary = _get_export_data(db, start_date, end_date, department_id)

    # Create Excel file
    excel_file = create_excel_file(student_list, summary)

    # Generate filename
    date_str = turkey_now().strftime("%Y%m%d_%H%M%S")
    filename = f"ogrenci_kayitlari_{date_str}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/daily/{date}")
async def export_daily(
    date: str,
    db: Session = Depends(get_db),
    current_user = Depends(require_admin)
):
    """Export student data for a specific date"""
    try:
        target_date = datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")

    start_date = target_date.replace(hour=0, minute=0, second=0, microsecond=0)
    end_date = start_date + timedelta(days=1)

    student_list, summary = _get_export_data(db, start_date, end_date, None)

    # Create Excel file
    excel_file = create_excel_file(student_list, summary)

    # Generate filename
    filename = f"ogrenci_kayitlari_{date}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
